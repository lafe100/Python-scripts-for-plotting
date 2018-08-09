# -*- coding: utf-8 -*-
"""
Created on Mon May 21 10:05:53 2018

@author: s1731217
"""

# =============================================================================
# model evaluation (obs, model, model (nudging)); NCP, YRD, PRD; NO2
# =============================================================================
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
import xlrd
from matplotlib.dates import DateFormatter, MonthLocator, WeekdayLocator, DayLocator, MONDAY, YEARLY
import matplotlib.dates as dates
from pylab import *
import numpy as np

# =============================================================================
# Reading data
# =============================================================================
#data = xlrd.open_workbook('/exports/csce/datastore/geos/users/s1731217/output_ukca/observation/beijing/Beijing.xlsx')  # 01/06/2014-31/10/2017
data = xlrd.open_workbook('C:/Users/s1731217/measurement data/CN Y2014/Nanjing/Nanjing.xlsx')
#table = data.sheets()[1] #search sheet for index
table1 = data.sheet_by_name('迈皋桥')
table2 = data.sheet_by_name('草场门')
table3 = data.sheet_by_name('山西路')
table4 = data.sheet_by_name('中华门')
table5 = data.sheet_by_name('瑞金路')
table6 = data.sheet_by_name('玄武湖')
table7 = data.sheet_by_name('浦口')
table8 = data.sheet_by_name('奥体中心')
table9 = data.sheet_by_name('仙林大学城')

col_time_old_1 = table1.col_values(0)[1:]    # col 1 for old missed time list
col_time_new_1 = table1.col_values(18)[1:]    # col 4 for new time list (using excel formula to fill in the missing data)
#  =S$1+ROW(S1)/24 in excel
col_values_old_1 = table1.col_values(10)[1:]  # col 3 for values

col_time_old_2 = table2.col_values(0)[1:]   
col_time_new_2 = table2.col_values(18)[1:]   
col_values_old_2 = table2.col_values(10)[1:]

col_time_old_3 = table3.col_values(0)[1:]   
col_time_new_3 = table3.col_values(18)[1:]   
col_values_old_3 = table3.col_values(10)[1:]

col_time_old_4 = table4.col_values(0)[1:]   
col_time_new_4 = table4.col_values(18)[1:]         
col_values_old_4 = table4.col_values(10)[1:]

col_time_old_5 = table5.col_values(0)[1:]   
col_time_new_5 = table5.col_values(18)[1:]   
col_values_old_5 = table5.col_values(10)[1:]

col_time_old_6 = table6.col_values(0)[1:]   
col_time_new_6 = table6.col_values(18)[1:]   
col_values_old_6 = table6.col_values(10)[1:]

col_time_old_7 = table7.col_values(0)[1:]   
col_time_new_7 = table7.col_values(18)[1:]   
col_values_old_7 = table7.col_values(10)[1:]

col_time_old_8 = table8.col_values(0)[1:]   
col_time_new_8 = table8.col_values(18)[1:]   
col_values_old_8 = table8.col_values(4)[1:]

col_time_old_9 = table9.col_values(0)[1:]   
col_time_new_9 = table9.col_values(18)[1:]   
col_values_old_9 = table9.col_values(4)[1:]

# =============================================================================
# Loading data
# =============================================================================
def hourly_mean(col_time_old_location, col_values_old_location, col_time_new_location):
    time_list_old = []
    length = 0
    for i in col_time_old_location:
        if i != '':
            length += 1
    print(length)
    for i in range(length):   # old time list 26817 values 
#    y = xlrd.xldate.xldate_as_datetime(table.cell(i, 0).value, 0)  
        y = xlrd.xldate.xldate_as_datetime(col_time_old_location[i], 0)  
        date_tmp = y.strftime('%Y/%m/%d %H:%M')
        time_list_old.append(date_tmp)

    time_list_new = []
    for i in range(len(col_time_new_1)):   # new time list 29976 values
#    y = xlrd.xldate.xldate_as_datetime(table.cell(i, 18).value, 0)
        y = xlrd.xldate.xldate_as_datetime(col_time_new_location[i], 0)  
        date_tmp = y.strftime('%Y/%m/%d %H:%M')
        time_list_new.append(date_tmp)

#time_list_final = []     # 为了写入X坐标轴能读取的数据
#for i in time_list_new:
#    zz = datetime.datetime.strptime(i, '%Y/%m/%d %H:%M')
#    time_list_final.append(zz)
    
    for i in range(len(col_values_old_location)):
        if col_values_old_location[i]=='None':
            col_values_old_location[i]=col_values_old_location[i-1]

    print(len(time_list_new))  # 29976
    print(len(time_list_old))  # 26817
    print(len(col_values_old_location))  # 29976

    test_value = []         # col_values_old_location has empty values
    for i in col_values_old_location:
        if i != '':
            test_value.append(i)
#print(test_value)
    print(len(test_value))  # 26817

# =============================================================================
# Filling the missing time values
# =============================================================================
    col_values_new = []
    
    for i in range(len(time_list_new)):
        col_values_new.append('none')  # 给新的列表全部赋值none

    indice = []                        # i为序号
    for i in time_list_old:
        indice.append(time_list_new.index(i))  # 取出旧的时间列表在新的时间列表中的序号

# for j in range(len(indice)): # 此种算法不能改变坏点时间点的顺序
#    col_values_new[indice[j]] = col_values_old_location[j] #赋值

    for i in indice:
        time = time_list_new[i]
        value = col_values_old_location[time_list_old.index(time)]
        col_values_new[i] = value

    for i in range(len(col_values_new)):  # 替换为none的值
        if col_values_new[i] == 'none':
            col_values_new[i] = col_values_new[i - 1]

    print(len(col_values_new))  # 29976
    print(len(time_list_new))  # 29976
#print(col_values_new)
    return(col_values_new)
    
df1 = pd.DataFrame(hourly_mean(col_time_old_1, col_values_old_1, col_time_new_1)) 
df2 = pd.DataFrame(hourly_mean(col_time_old_2, col_values_old_2, col_time_new_2)) 
df3 = pd.DataFrame(hourly_mean(col_time_old_3, col_values_old_3, col_time_new_3)) 
df4 = pd.DataFrame(hourly_mean(col_time_old_4, col_values_old_4, col_time_new_4)) 
df5 = pd.DataFrame(hourly_mean(col_time_old_5, col_values_old_5, col_time_new_5)) 
df6 = pd.DataFrame(hourly_mean(col_time_old_6, col_values_old_6, col_time_new_6)) 
df7 = pd.DataFrame(hourly_mean(col_time_old_7, col_values_old_7, col_time_new_7))
df8 = pd.DataFrame(hourly_mean(col_time_old_8, col_values_old_8, col_time_new_8)) 
df9 = pd.DataFrame(hourly_mean(col_time_old_9, col_values_old_9, col_time_new_9))

writer = pd.ExcelWriter('C:/Users/s1731217/measurement data/CN Y2014/Nanjing/no2_obs_nanjing_hourly.xlsx')
df1.to_excel(writer, 'AVERAGE', index = False, header = ['迈皋桥'], startcol = 0, merge_cells=False)
df2.to_excel(writer, 'AVERAGE', index = False, header = ['草场门'], startcol = 1, merge_cells=False)
df3.to_excel(writer, 'AVERAGE', index = False, header = ['山西路'], startcol = 2, merge_cells=False)
df4.to_excel(writer, 'AVERAGE', index = False, header = ['中华门'], startcol = 3, merge_cells=False)
df5.to_excel(writer, 'AVERAGE', index = False, header = ['瑞金路'], startcol = 4, merge_cells=False)
df6.to_excel(writer, 'AVERAGE', index = False, header = ['玄武湖'], startcol = 5, merge_cells=False)
df7.to_excel(writer, 'AVERAGE', index = False, header = ['浦口'],   startcol = 6, merge_cells=False)
df8.to_excel(writer, 'AVERAGE', index = False, header = ['奥体中心'], startcol = 7, merge_cells=False)
df9.to_excel(writer, 'AVERAGE', index = False, header = ['仙林大学城'], startcol = 8, merge_cells=False)

writer.save()
